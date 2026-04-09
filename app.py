from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import sqlite3, os, json
from datetime import datetime
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'sbhc-secret-change-in-prod')
DB = 'dashboard.db'
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ─── DB ───────────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as db:
        db.executescript('''
            CREATE TABLE IF NOT EXISTS users (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                name      TEXT NOT NULL,
                email     TEXT UNIQUE NOT NULL,
                password  TEXT NOT NULL,
                company   TEXT DEFAULT '',
                role      TEXT DEFAULT 'admin',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS ventas (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id         INTEGER NOT NULL,
                fecha           TEXT,
                producto        TEXT,
                ingresos        REAL DEFAULT 0,
                gastos          REAL DEFAULT 0,
                clientes_nuevos INTEGER DEFAULT 0,
                canal           TEXT DEFAULT 'Directo',
                FOREIGN KEY (user_id) REFERENCES users(id)
            );
        ''')

init_db()

# ─── AUTH DECORATOR ───────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

# ─── EXCEL PARSER ─────────────────────────────────────────────────────────────

REQUIRED_COLS = {'fecha', 'producto', 'ingresos', 'gastos', 'clientes_nuevos', 'canal'}

def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El archivo está vacío.")
    headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[0]]
    missing = REQUIRED_COLS - set(headers)
    if missing:
        raise ValueError(f"Faltan columnas: {', '.join(sorted(missing))}.")
    data = []
    for i, row in enumerate(rows[1:], start=2):
        if all(v is None for v in row):
            continue
        d = dict(zip(headers, row))
        try:
            ingresos = float(d.get('ingresos') or 0)
            gastos   = float(d.get('gastos')   or 0)
            clientes = int(float(d.get('clientes_nuevos') or 0))
        except (ValueError, TypeError):
            raise ValueError(f"Fila {i}: ingresos, gastos y clientes_nuevos deben ser números.")
        # Fecha: puede ser datetime (Excel) o string
        raw_fecha = d.get('fecha')
        if isinstance(raw_fecha, datetime):
            fecha = raw_fecha.strftime('%Y-%m-%d')
        else:
            fecha = str(raw_fecha or '').strip()[:10]
        data.append({
            'fecha':           fecha,
            'producto':        str(d.get('producto') or 'Sin producto').strip(),
            'ingresos':        ingresos,
            'gastos':          gastos,
            'clientes_nuevos': clientes,
            'canal':           str(d.get('canal') or 'Directo').strip(),
        })
    if not data:
        raise ValueError("No hay filas de datos (solo cabecera).")
    return data

# ─── KPI CALCULATOR (datos reales) ────────────────────────────────────────────

def fmt_eur(n):
    if n >= 1_000_000: return f"€{n/1_000_000:.2f}M"
    if n >= 1_000:     return f"€{n/1_000:.1f}k"
    return f"€{round(n):,}"

def calculate_kpis_from_rows(rows):
    MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    total_rev     = sum(r['ingresos'] for r in rows)
    total_exp     = sum(r['gastos']   for r in rows)
    total_profit  = total_rev - total_exp
    beneficio_pct = round(total_profit / total_rev * 100, 1) if total_rev else 0
    total_clientes = sum(r['clientes_nuevos'] for r in rows)

    # Por mes
    by_month = defaultdict(lambda: {'ingresos': 0, 'gastos': 0})
    for r in rows:
        try:
            dt  = datetime.strptime(r['fecha'][:7], '%Y-%m')
            key = (dt.year, dt.month)
        except Exception:
            key = (0, 0)
        by_month[key]['ingresos'] += r['ingresos']
        by_month[key]['gastos']   += r['gastos']

    sorted_months = sorted(k for k in by_month if k != (0, 0))[-6:]
    labels   = [MESES[m-1] for _, m in sorted_months] if sorted_months else ['—']
    revenue  = [round(by_month[k]['ingresos'] / 1000, 2) for k in sorted_months]
    expenses = [round(by_month[k]['gastos']   / 1000, 2) for k in sorted_months]
    profit   = [round(r - e, 2) for r, e in zip(revenue, expenses)]

    # Canales
    by_canal = defaultdict(float)
    for r in rows:
        by_canal[r['canal']] += r['ingresos']
    channels = {k: round(v) for k, v in sorted(by_canal.items(), key=lambda x: -x[1])}

    # Productos
    by_prod = defaultdict(lambda: {'ventas': 0, 'ingresos': 0, 'gastos': 0})
    for r in rows:
        by_prod[r['producto']]['ventas']   += r['clientes_nuevos']
        by_prod[r['producto']]['ingresos'] += r['ingresos']
        by_prod[r['producto']]['gastos']   += r['gastos']

    products = []
    for nombre, d in sorted(by_prod.items(), key=lambda x: -x[1]['ingresos'])[:6]:
        ing = d['ingresos']
        gas = d['gastos']
        margen = round((ing - gas) / ing * 100) if ing else 0
        products.append({'name': nombre, 'ventas': d['ventas'], 'ingresos': round(ing), 'margen': margen})

    n_meses = max(len(sorted_months), 1)
    mrr = round(total_rev / n_meses)
    arr = mrr * 12
    cac = round(total_exp / total_clientes) if total_clientes else 0
    ltv = round(mrr * 24 / total_clientes)  if total_clientes else 0
    ltv_cac    = round(ltv / cac, 1) if cac else 0
    conversion = round(total_clientes / max(len(rows), 1) * 100, 1)

    goals = [
        {"nombre": "Ingresos objetivo",   "objetivo": round(total_rev * 1.2),       "actual": round(total_rev),    "color": "#2dd4bf"},
        {"nombre": "Clientes nuevos",      "objetivo": round(total_clientes * 1.3),  "actual": total_clientes,      "color": "#4ade80"},
        {"nombre": "Margen neto %",        "objetivo": 40,                            "actual": beneficio_pct,       "color": "#60a5fa"},
        {"nombre": "Conversión %",         "objetivo": 10,                            "actual": conversion,          "color": "#fbbf24"},
    ]

    activity = [
        {"evento": f"Excel cargado — {len(rows)} transacciones", "color": "green",  "hace": "ahora"},
        {"evento": f"Producto estrella: {products[0]['name']}" if products else "Sin productos", "color": "purple", "hace": "calculado"},
        {"evento": f"Canal líder: {list(channels.keys())[0]}" if channels else "Sin canales", "color": "blue", "hace": "calculado"},
        {"evento": f"MRR: {fmt_eur(mrr)} · ARR: {fmt_eur(arr)}", "color": "green",  "hace": "calculado"},
        {"evento": f"CAC: {fmt_eur(cac)} · LTV: {fmt_eur(ltv)}", "color": "amber",  "hace": "calculado"},
    ]

    return {
        "source": "excel",
        "kpis": {
            "ingresos_total": fmt_eur(total_rev),
            "ingresos_raw":   round(total_rev / 1000, 1),
            "gastos_total":   fmt_eur(total_exp),
            "beneficio":      fmt_eur(total_profit),
            "beneficio_pct":  beneficio_pct,
            "clientes":       total_clientes,
            "churn":          "—",
            "nps":            "—",
            "conversion":     f"{conversion}%",
            "mrr":            fmt_eur(mrr),
            "arr":            fmt_eur(arr),
            "cac":            fmt_eur(cac),
            "ltv":            fmt_eur(ltv),
            "ltv_cac":        ltv_cac,
        },
        "charts":   {"labels": labels, "revenue": revenue, "expenses": expenses, "profit": profit, "channels": channels},
        "products": products,
        "goals":    goals,
        "activity": activity,
    }

# ─── ESTADO VACÍO (cuenta nueva, sin datos) ──────────────────────────────────

def empty_kpis():
    return {
        "source": "empty",
        "kpis": {
            "ingresos_total": "—", "ingresos_raw": 0,
            "gastos_total": "—", "beneficio": "—", "beneficio_pct": 0,
            "clientes": 0, "churn": "—", "nps": "—",
            "conversion": "—", "mrr": "—", "arr": "—",
            "cac": "—", "ltv": "—", "ltv_cac": 0,
        },
        "charts": {"labels": [], "revenue": [], "expenses": [], "profit": [], "channels": {}},
        "products": [],
        "goals": [],
        "activity": [{"evento": "Bienvenido — importa tu Excel para empezar", "color": "blue", "hace": "ahora"}],
    }

# ─── ROUTES ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user_id' in session else url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))

    login_error = register_error = register_success = None
    show_register = request.args.get('tab') == 'register'

    if request.method == 'POST':
        tab = request.form.get('tab', 'login')

        if tab == 'login':
            email    = request.form.get('email', '').strip().lower()
            password = request.form.get('password', '')
            db   = get_db()
            user = db.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
            if user and check_password_hash(user['password'], password):
                session['user_id']      = user['id']
                session['user_name']    = user['name']
                session['user_company'] = user['company'] or ''
                session['user_email']   = user['email']
                return redirect(url_for('dashboard'))
            login_error = 'Email o contraseña incorrectos.'

        elif tab == 'register':
            show_register = True
            name     = request.form.get('name',         '').strip()
            company  = request.form.get('company',      '').strip()
            email    = request.form.get('reg_email',    '').strip().lower()
            password = request.form.get('reg_password', '')
            confirm  = request.form.get('reg_confirm',  '')

            if not name or not email or not password:
                register_error = 'Nombre, email y contraseña son obligatorios.'
            elif password != confirm:
                register_error = 'Las contraseñas no coinciden.'
            elif len(password) < 6:
                register_error = 'La contraseña debe tener al menos 6 caracteres.'
            else:
                try:
                    db = get_db()
                    db.execute(
                        'INSERT INTO users (name, email, company, password) VALUES (?, ?, ?, ?)',
                        (name, email, company, generate_password_hash(password))
                    )
                    db.commit()
                    user = db.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
                    session['user_id']      = user['id']
                    session['user_name']    = user['name']
                    session['user_company'] = company
                    session['user_email']   = email
                    return redirect(url_for('dashboard'))
                except sqlite3.IntegrityError:
                    register_error = 'Este email ya está registrado. Inicia sesión.'

    return render_template('login.html',
        login_error=login_error,
        register_error=register_error,
        register_success=register_success,
        show_register=show_register,
    )

@app.route('/register')
def register():
    return redirect(url_for('login') + '?tab=register')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    period = request.args.get('period', 'Q3 2025')
    db   = get_db()
    rows = db.execute('SELECT * FROM ventas WHERE user_id = ?', (session['user_id'],)).fetchall()
    if rows:
        data = calculate_kpis_from_rows([dict(r) for r in rows])
    else:
        data = empty_kpis()
    return render_template('dashboard.html',
        user_name=session['user_name'],
        user_company=session.get('user_company', ''),
        user_email=session['user_email'],
        period=period,
        data=data,
        data_json=json.dumps(data),
    )

@app.route('/importar', methods=['GET', 'POST'])
@login_required
def importar():
    error = success = preview = stats = None

    if request.method == 'POST':
        action = request.form.get('action', 'upload')

        if action == 'clear':
            db = get_db()
            db.execute('DELETE FROM ventas WHERE user_id = ?', (session['user_id'],))
            db.commit()
            success = 'Datos eliminados correctamente. El dashboard está en blanco.'

        else:  # upload
            file = request.files.get('excel_file')
            if not file or file.filename == '':
                error = 'Selecciona un archivo Excel primero.'
            elif not file.filename.lower().endswith(('.xlsx', '.xlsm')):
                error = 'Solo se aceptan archivos .xlsx o .xlsm'
            else:
                filename = secure_filename(f"u{session['user_id']}_{file.filename}")
                filepath = os.path.join(UPLOAD_FOLDER, filename)
                file.save(filepath)
                try:
                    rows_data = parse_excel(filepath)
                    db = get_db()
                    db.execute('DELETE FROM ventas WHERE user_id = ?', (session['user_id'],))
                    db.executemany(
                        'INSERT INTO ventas (user_id,fecha,producto,ingresos,gastos,clientes_nuevos,canal) VALUES (?,?,?,?,?,?,?)',
                        [(session['user_id'], r['fecha'], r['producto'], r['ingresos'], r['gastos'], r['clientes_nuevos'], r['canal']) for r in rows_data]
                    )
                    db.commit()
                    os.remove(filepath)
                    preview = rows_data[:5]
                    stats = {
                        'filas':    len(rows_data),
                        'ingresos': round(sum(r['ingresos'] for r in rows_data), 2),
                        'gastos':   round(sum(r['gastos']   for r in rows_data), 2),
                        'clientes': sum(r['clientes_nuevos'] for r in rows_data),
                        'productos': len(set(r['producto'] for r in rows_data)),
                        'canales':  len(set(r['canal']    for r in rows_data)),
                    }
                    success = f'¡Importación correcta! {len(rows_data)} filas cargadas.'
                except ValueError as e:
                    error = str(e)
                    if os.path.exists(filepath): os.remove(filepath)
                except Exception as e:
                    error = f'Error inesperado al leer el archivo: {str(e)}'
                    if os.path.exists(filepath): os.remove(filepath)

    db    = get_db()
    count = db.execute('SELECT COUNT(*) FROM ventas WHERE user_id = ?', (session['user_id'],)).fetchone()[0]
    return render_template('importar.html',
        user_name=session['user_name'],
        user_company=session.get('user_company', ''),
        error=error, success=success, preview=preview, stats=stats,
        rows_count=count,
    )

@app.route('/api/kpis')
@login_required
def api_kpis():
    db   = get_db()
    rows = db.execute('SELECT * FROM ventas WHERE user_id = ?', (session['user_id'],)).fetchall()
    data = calculate_kpis_from_rows([dict(r) for r in rows]) if rows else empty_kpis()
    return jsonify(data)

@app.route('/descargar_plantilla')
@login_required
def descargar_plantilla():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    headers = ['fecha', 'producto', 'ingresos', 'gastos', 'clientes_nuevos', 'canal']
    header_fill = PatternFill(start_color="1A2535", end_color="1A2535", fill_type="solid")
    header_font = Font(bold=True, color="2DD4BF", name="Calibri")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 20
    ejemplos = [
        ['2025-01-08','Plan Básico',350,120,5,'Orgánico'],
        ['2025-01-15','Plan Pro',1200,400,3,'Directo'],
        ['2025-01-22','Plan Enterprise',4500,800,1,'Referidos'],
        ['2025-02-05','Consultoría',2000,600,2,'Directo'],
        ['2025-02-12','Plan Básico',700,240,8,'Publicidad'],
    ]
    for row_data in ejemplos:
        ws.append(row_data)
    path = os.path.join(UPLOAD_FOLDER, f'plantilla_{session["user_id"]}.xlsx')
    wb.save(path)
    return send_file(path, as_attachment=True, download_name='plantilla_sbhc.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    db   = get_db()
    user = db.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if user is None:
        session.clear()
        return redirect(url_for('login'))
    success = error = None
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'update':
            name    = request.form.get('name', '').strip()
            company = request.form.get('company', '').strip()
            if not name:
                error = 'El nombre no puede estar vacío.'
            else:
                db.execute('UPDATE users SET name=?, company=? WHERE id=?', (name, company, session['user_id']))
                db.commit()
                session['user_name']    = name
                session['user_company'] = company
                success = 'Perfil actualizado correctamente.'
                user = db.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        elif action == 'password':
            current  = request.form.get('current_password', '')
            new_pass = request.form.get('new_password', '')
            confirm  = request.form.get('confirm_password', '')
            if not check_password_hash(user['password'], current):
                error = 'La contraseña actual es incorrecta.'
            elif new_pass != confirm:
                error = 'Las contraseñas nuevas no coinciden.'
            elif len(new_pass) < 6:
                error = 'La contraseña debe tener al menos 6 caracteres.'
            else:
                db.execute('UPDATE users SET password=? WHERE id=?',
                           (generate_password_hash(new_pass), session['user_id']))
                db.commit()
                success = 'Contraseña cambiada correctamente.'
    return render_template('profile.html',
        user=user,
        user_name=session['user_name'],
        user_company=session.get('user_company', ''),
        success=success, error=error,
    )


# ─── ANALÍTICAS ───────────────────────────────────────────────────────────────

def get_analiticas_data(user_id):
    """Calcula datos detallados para la página de analíticas."""
    db   = get_db()
    rows = db.execute('SELECT * FROM ventas WHERE user_id = ?', (user_id,)).fetchall()
    if not rows:
        return None
    rows = [dict(r) for r in rows]
    MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']

    # Por mes (todos los meses disponibles)
    by_month = defaultdict(lambda: {'ingresos':0,'gastos':0,'clientes':0,'transacciones':0})
    for r in rows:
        try:
            dt  = datetime.strptime(r['fecha'][:7], '%Y-%m')
            key = (dt.year, dt.month)
        except:
            continue
        by_month[key]['ingresos']      += r['ingresos']
        by_month[key]['gastos']        += r['gastos']
        by_month[key]['clientes']      += r['clientes_nuevos']
        by_month[key]['transacciones'] += 1

    sorted_months = sorted(by_month.keys())
    labels_mes    = [f"{MESES[m-1]} {y}" for y, m in sorted_months]
    rev_mes       = [round(by_month[k]['ingresos'], 2)  for k in sorted_months]
    gas_mes       = [round(by_month[k]['gastos'], 2)    for k in sorted_months]
    ben_mes       = [round(by_month[k]['ingresos'] - by_month[k]['gastos'], 2) for k in sorted_months]
    cli_mes       = [by_month[k]['clientes']            for k in sorted_months]
    txn_mes       = [by_month[k]['transacciones']       for k in sorted_months]
    margen_mes    = [round((b/r*100),1) if r else 0 for b,r in zip(ben_mes, rev_mes)]

    # Por canal
    by_canal = defaultdict(lambda: {'ingresos':0,'clientes':0,'txn':0})
    for r in rows:
        c = r['canal'] or 'Sin canal'
        by_canal[c]['ingresos']  += r['ingresos']
        by_canal[c]['clientes']  += r['clientes_nuevos']
        by_canal[c]['txn']       += 1
    canal_sorted = sorted(by_canal.items(), key=lambda x: -x[1]['ingresos'])
    canal_labels  = [c for c,_ in canal_sorted]
    canal_ing     = [round(d['ingresos'],2) for _,d in canal_sorted]
    canal_cli     = [d['clientes'] for _,d in canal_sorted]

    # Por producto
    by_prod = defaultdict(lambda: {'ingresos':0,'gastos':0,'clientes':0,'txn':0})
    for r in rows:
        p = r['producto'] or 'Sin producto'
        by_prod[p]['ingresos'] += r['ingresos']
        by_prod[p]['gastos']   += r['gastos']
        by_prod[p]['clientes'] += r['clientes_nuevos']
        by_prod[p]['txn']      += 1
    prod_sorted = sorted(by_prod.items(), key=lambda x: -x[1]['ingresos'])
    prod_labels  = [p for p,_ in prod_sorted]
    prod_ing     = [round(d['ingresos'],2) for _,d in prod_sorted]
    prod_margen  = [round((d['ingresos']-d['gastos'])/d['ingresos']*100,1) if d['ingresos'] else 0 for _,d in prod_sorted]

    # Tabla de transacciones (últimas 20)
    txn_table = sorted(rows, key=lambda x: x.get('fecha',''), reverse=True)[:20]

    # Resumen global
    total_rev = sum(r['ingresos'] for r in rows)
    total_exp = sum(r['gastos']   for r in rows)
    total_cli = sum(r['clientes_nuevos'] for r in rows)
    n_meses   = max(len(sorted_months), 1)

    return {
        'tiene_datos': True,
        'resumen': {
            'total_ingresos': fmt_eur(total_rev),
            'total_gastos':   fmt_eur(total_exp),
            'total_beneficio':fmt_eur(total_rev - total_exp),
            'margen_pct':     round((total_rev-total_exp)/total_rev*100,1) if total_rev else 0,
            'total_clientes': total_cli,
            'total_txn':      len(rows),
            'mejor_mes':      labels_mes[rev_mes.index(max(rev_mes))] if rev_mes else '—',
            'mejor_producto': prod_labels[0] if prod_labels else '—',
            'mejor_canal':    canal_labels[0] if canal_labels else '—',
            'ticket_medio':   fmt_eur(total_rev / len(rows)) if rows else '—',
        },
        'por_mes': {
            'labels': labels_mes,
            'ingresos': rev_mes,
            'gastos':   gas_mes,
            'beneficio':ben_mes,
            'clientes': cli_mes,
            'margen':   margen_mes,
            'txn':      txn_mes,
        },
        'por_canal': {
            'labels':   canal_labels,
            'ingresos': canal_ing,
            'clientes': canal_cli,
        },
        'por_producto': {
            'labels': prod_labels,
            'ingresos': prod_ing,
            'margen':   prod_margen,
        },
        'txn_recientes': txn_table,
    }

@app.route('/analiticas')
@login_required
def analiticas():
    data = get_analiticas_data(session['user_id'])
    return render_template('analiticas.html',
        user_name=session['user_name'],
        user_company=session.get('user_company',''),
        data=data,
        data_json=json.dumps(data) if data else 'null',
    )

# ─── OBJETIVOS ────────────────────────────────────────────────────────────────

def init_objetivos_table():
    with get_db() as db:
        db.execute("""
            CREATE TABLE IF NOT EXISTS objetivos (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id    INTEGER NOT NULL,
                nombre     TEXT NOT NULL,
                tipo       TEXT NOT NULL,
                objetivo   REAL NOT NULL,
                color      TEXT DEFAULT '#2dd4bf',
                activo     INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
        """)

init_objetivos_table()

TIPO_LABELS = {
    'ingresos_total': 'Ingresos totales (€)',
    'beneficio_pct':  'Margen de beneficio (%)',
    'clientes_total': 'Clientes nuevos',
    'ticket_medio':   'Ticket medio (€)',
    'txn_total':      'Nº de transacciones',
}

def calcular_actual(tipo, rows):
    if not rows:
        return 0
    total_rev = sum(r['ingresos'] for r in rows)
    total_exp = sum(r['gastos']   for r in rows)
    total_cli = sum(r['clientes_nuevos'] for r in rows)
    if tipo == 'ingresos_total': return round(total_rev, 2)
    if tipo == 'beneficio_pct':  return round((total_rev-total_exp)/total_rev*100,1) if total_rev else 0
    if tipo == 'clientes_total': return total_cli
    if tipo == 'ticket_medio':   return round(total_rev/len(rows),2) if rows else 0
    if tipo == 'txn_total':      return len(rows)
    return 0

@app.route('/objetivos', methods=['GET','POST'])
@login_required
def objetivos():
    init_objetivos_table()
    db   = get_db()
    error = success = None

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add':
            nombre   = request.form.get('nombre','').strip()
            tipo     = request.form.get('tipo','')
            objetivo = request.form.get('objetivo','').strip()
            color    = request.form.get('color','#2dd4bf')
            if not nombre or not tipo or not objetivo:
                error = 'Todos los campos son obligatorios.'
            else:
                try:
                    obj_val = float(objetivo)
                    db.execute(
                        'INSERT INTO objetivos (user_id,nombre,tipo,objetivo,color) VALUES (?,?,?,?,?)',
                        (session['user_id'], nombre, tipo, obj_val, color)
                    )
                    db.commit()
                    success = f'Objetivo "{nombre}" creado correctamente.'
                except ValueError:
                    error = 'El valor del objetivo debe ser un número.'

        elif action == 'delete':
            obj_id = request.form.get('obj_id')
            db.execute('DELETE FROM objetivos WHERE id=? AND user_id=?', (obj_id, session['user_id']))
            db.commit()
            success = 'Objetivo eliminado.'

        elif action == 'toggle':
            obj_id = request.form.get('obj_id')
            db.execute('UPDATE objetivos SET activo = 1 - activo WHERE id=? AND user_id=?', (obj_id, session['user_id']))
            db.commit()

    # Cargar objetivos del usuario
    objs = db.execute('SELECT * FROM objetivos WHERE user_id=? ORDER BY created_at DESC', (session['user_id'],)).fetchall()
    rows = db.execute('SELECT * FROM ventas WHERE user_id=?', (session['user_id'],)).fetchall()
    rows_list = [dict(r) for r in rows]

    # Calcular progreso de cada objetivo
    objetivos_data = []
    for o in objs:
        actual = calcular_actual(o['tipo'], rows_list)
        pct    = min(round(actual / o['objetivo'] * 100, 1), 200) if o['objetivo'] else 0
        pct_display = min(pct, 100)
        objetivos_data.append({
            'id':       o['id'],
            'nombre':   o['nombre'],
            'tipo':     o['tipo'],
            'tipo_label': TIPO_LABELS.get(o['tipo'], o['tipo']),
            'objetivo': o['objetivo'],
            'actual':   actual,
            'pct':      pct,
            'pct_display': pct_display,
            'color':    o['color'],
            'activo':   o['activo'],
            'superado': pct >= 100,
        })

    tiene_datos = len(rows_list) > 0

    return render_template('objetivos.html',
        user_name=session['user_name'],
        user_company=session.get('user_company',''),
        objetivos=objetivos_data,
        tipos=TIPO_LABELS,
        error=error,
        success=success,
        tiene_datos=tiene_datos,
        total_objs=len(objetivos_data),
        superados=sum(1 for o in objetivos_data if o['superado']),
    )

if __name__ == '__main__':
    app.run(debug=True, port=5000)
